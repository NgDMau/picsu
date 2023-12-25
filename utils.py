import openpyxl
import random
import django
import os

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'picsu.settings')  # Replace 'your_project.settings' with your actual Django project's settings
django.setup()

from study.models import Word, Answer, Question

def import_data(excel_path, learning_method):
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    
    if learning_method == 'picsu':
        print("insert picsu")
        all_words = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming first row is header
            word_text, synonym1, synonym2, synonym3, chinese_meaning, english_meaning, image_url = row

            # Create Word instance
            word, created = Word.objects.get_or_create(original_word=word_text, english_meaning=english_meaning, chinese_meaning=chinese_meaning, image_url=image_url, learning_method='picsu')
            all_words.append(word)
            # Create Question instance
            question = Question.objects.create(word=word, group='p1')
            # Create Answer instances and link to Question
            for synonym in [synonym1, synonym2, synonym3]:
                if synonym:  # Skip blank synonyms
                    answer = Answer.objects.create(word=word, text=synonym)
                    question.correct_answers.add(answer)
            

        random.shuffle(all_words)
        p1, p2 = all_words[0:60], all_words[60:120]
        print("p1 p2 ", len(p1), len(p2))

        # p1 = random.sample(all_words, min(80, len(all_words)))
        # p2 = random.sample(all_words, min(80, len(all_words)))

        for word in p1:
            if not Question.objects.filter(word=word).exists():
                print("Word for p1 doesnt exist!")
            else:
                Question.objects.filter(word=word).update(group='p1')
                print("Inserted for p1")
        for word in p2:
            if not Question.objects.filter(word=word).exists():
                print("Word for p2 doesnt exist!")
            else:
                Question.objects.filter(word=word).update(group='p2')
                print("Inserted for p2")

        print("LEN Q P1-P2: ", len(Question.objects.filter(group='p1')), len(Question.objects.filter(group='p2')))

    elif learning_method == 'flashcard':
        print("insert flashcard")
        all_words = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming first row is header
            word_text, synonym1, synonym2, synonym3, chinese_meaning, english_meaning = row

            # Create Word instance
            word, created = Word.objects.get_or_create(original_word=word_text, english_meaning=english_meaning, chinese_meaning=chinese_meaning, learning_method='flashcard')

            # Create Question instance
            question = Question.objects.create(word=word)

            all_words.append(word)

            # Create Answer instances and link to Question
            for synonym in [synonym1, synonym2, synonym3]:
                if synonym:  # Skip blank synonyms
                    answer = Answer.objects.create(word=word, text=synonym)
                    question.correct_answers.add(answer)

        # Divide words into two groups
        random.shuffle(all_words)
        halfway_point = len(all_words) // 2
        
        f1, f2 = all_words[:80], all_words[80:160]

        # Assign groups
        for word in f1:
            Question.objects.filter(word=word).update(group='f1')
            print("F1")
        for word in f2:
            Question.objects.filter(word=word).update(group='f2')
            print("F2")
        
        print("LEN Q P1-P2: ", len(Question.objects.filter(group='f1')), len(Question.objects.filter(group='f2')))




if __name__ == "__main__":
    picsu_dict_path = 'Vocabulary_PICSU.xlsx' 
    flashcard_dict_path = 'Vocabulary_Flashcard.xlsx'
    print("Running...")
    import_data(picsu_dict_path, "picsu")
    # import_data(flashcard_dict_path, "flashcard")
